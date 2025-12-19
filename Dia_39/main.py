# ========================================== #
# Manipulação de arquivos Excel com openpyxl #
# Leia e escreva dados em um arquivo Excel.  #
# ========================================== #

from openpyxl import Workbook, load_workbook
import os


ARQUIVO = "Dia_39/dados.xlsx"


def criar_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Pessoas"

    # Cabeçalhos
    ws.append(["Nome", "Idade", "Salário"])

    # Dados
    ws.append(["Ana", 30, 2500])
    ws.append(["João", 25, 3200])
    ws.append(["Maria", 40, 4100])

    wb.save(ARQUIVO)
    print("Arquivo Excel criado com sucesso!")


def ler_excel():
    wb = load_workbook(ARQUIVO)
    ws = wb.active

    for linha in ws.iter_rows(values_only=True):
        print(linha)


def escrever_excel(nome, idade, salario):
    wb = load_workbook(ARQUIVO)
    ws = wb.active

    ws.append([nome, idade, salario])
    wb.save(ARQUIVO)
    print("Dados adicionados ao Excel!")

if not os.path.exists(ARQUIVO):
    criar_excel()

print("\nDados atuais:")
ler_excel()

print("\nAdicionando novo registro...")
escrever_excel("Carlos", 28, 3500)

print("\nDados atualizados:")
ler_excel()

# ================================= #
# Para rodar: python -m Dia_39.main #
# ================================= #
